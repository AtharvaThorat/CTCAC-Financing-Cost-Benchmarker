import os
import re
import warnings
import requests
import pandas as pd
import openpyxl
from bs4 import BeautifulSoup
from urllib.parse import urljoin

# ignore openpyxl warnings about conditional formatting
warnings.filterwarnings("ignore")

# Config
BASE_URL = "https://www.treasurer.ca.gov/ctcac/2025/thirdround/4percent/application/index.asp"
DATA_DIR = "Downloaded files"
OUTPUT_FILE = "summary_output.csv"
TOLERANCE = 1.0

def download_files():
    if not os.path.exists(DATA_DIR):
        os.makedirs(DATA_DIR)
        
    print(f"Checking {BASE_URL}...")
    
    try:
        r = requests.get(BASE_URL)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, 'html.parser')
        
        # get all excel links
        links = [l for l in soup.find_all('a', href=True) 
                 if l['href'].lower().endswith(('.xlsx', '.xlsm', '.xls'))]
        
        print(f"Found {len(links)} files.")
        
        for i, link in enumerate(links, 1):
            fname = os.path.basename(link['href'])
            fname = requests.utils.unquote(fname)
            fpath = os.path.join(DATA_DIR, fname)
            full_url = urljoin(BASE_URL, link['href'])
            
            if not os.path.exists(fpath):
                print(f"[{i}/{len(links)}] Downloading {fname}...")
                try:
                    f_resp = requests.get(full_url)
                    f_resp.raise_for_status()
                    with open(fpath, 'wb') as f:
                        f.write(f_resp.content)
                except Exception as e:
                    print(f"Error downloading {fname}: {e}")
            else:
                print(f"[{i}/{len(links)}] Skipping {fname} (exists)")
                
        return True
                
    except Exception as e:
        print(f"Connection error: {e}")
        return False

# cleaning helpers
def clean_value(val):
    if val is None: return 0.0
    if isinstance(val, (int, float)): return float(val)
    
    if isinstance(val, str):
        val = val.strip()
        # handle (100) -> -100
        if val.startswith('(') and val.endswith(')'):
            val = '-' + val[1:-1]
        
        val = val.replace('$', '').replace(',', '')
        
        match = re.search(r'-?\d+(\.\d+)?', val)
        if match:
            try:
                return float(match.group())
            except:
                return 0.0
    return 0.0

def clean_text(val):
    if not val: return ""
    val = str(val).strip()
    val = re.sub(r'^Other[:\s-]*', '', val, flags=re.IGNORECASE)
    val = re.sub(r'[\(\)]', '', val)
    return val.strip()

# scans for first match in grid
def grid_scanner(sheet, keywords, min_val, max_val):
    rows = list(sheet.iter_rows(values_only=True))
    for i, row in enumerate(rows):
        row_str = " ".join([str(x).lower() for x in row if x])
        
        match = False
        for group in keywords:
            if all(k.lower() in row_str for k in group):
                match = True
                break
        
        if match:
            # check this row and next 2 for valid number
            for r_off in range(3):
                if i + r_off >= len(rows): break
                for cell in rows[i+r_off]:
                    val = clean_value(cell)
                    if min_val <= val <= max_val:
                        return val
    return 0

# global search for max SF
def find_sf_in_workbook(wb):
    candidates = []
    sheets = [s for s in wb.sheetnames if "App" in s] + \
             [s for s in wb.sheetnames if "App" not in s]
    
    for s_name in sheets:
        if any(x in s_name for x in ["Source", "Budget", "Cost", "Financ"]):
            continue
            
        ws = wb[s_name]
        for row in ws.iter_rows(values_only=True):
            row_str = " ".join([str(x).lower() for x in row if x])
            if any(k in row_str for k in ["sq. ft.", "square footage", "net rentable", "gross building", "gba", "residential area"]):
                for cell in row:
                    val = clean_value(cell)
                    if 2000 <= val <= 2000000:
                        candidates.append(val)
    
    return max(candidates) if candidates else 0

def parse_application(fpath):
    fname = os.path.basename(fpath)
    
    row = {
        "File Name": fname, 
        "Flag": [], 
        "Combined Financing Costs": 0.0, 
        "Financing Cost per Unit": 0.0, 
        "Financing Cost per SF": 0.0, 
        "% of Hard Costs": 0.0,
        "Total Units": 0, "Total SF": 0, "Hard Costs": 0.0,
        
        "Const Loan Interest": 0.0, "Const Origination Fee": 0.0,
        "Const Credit Enhancement": 0.0, "Const Bond Premium": 0.0,
        "Const Cost of Issuance": 0.0, "Const Title & Recording": 0.0,
        "Const Taxes": 0.0, "Const Insurance": 0.0,
        "Const Other Costs": 0.0, "Const Other Details": [],
        "Const Total (Calculated)": 0.0, "Const Total (Sheet)": 0.0,
        
        "Perm Loan Origination Fee": 0.0, "Perm Credit Enhancement": 0.0,
        "Perm Title & Recording": 0.0, "Perm Taxes": 0.0, "Perm Insurance": 0.0,
        "Perm Other Costs": 0.0, "Perm Other Details": [],
        "Perm Total (Calculated)": 0.0, "Perm Total (Sheet)": 0.0
    }

    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
    except Exception as e:
        row["Flag"].append(f"Error: {e}")
        return flatten_output(row)

    # 1. Sources & Uses parsing
    su_name = next((s for s in wb.sheetnames if "Sources" in s and ("Uses" in s or "Budget" in s)), None)
    if not su_name: 
        su_name = next((s for s in wb.sheetnames if "S&U" in s), None)

    if su_name:
        ws = wb[su_name]

        def get_header(sheet, kws):
            for r in sheet.iter_rows():
                for c in r:
                    if c.value and isinstance(c.value, str):
                        for k in kws:
                            if k.lower() in c.value.lower(): return c
            return None

        def get_section_data(h_kw, t_kw, mapping, prefix):
            h_cell = get_header(ws, h_kw)
            if not h_cell: return
            
            curr = h_cell.row + 1
            for _ in range(40):
                lbl_cell = ws.cell(row=curr, column=h_cell.column)
                # values usually col 18 or 3
                val_col = 18 if ws.max_column >= 18 else 3
                val_cell = ws.cell(row=curr, column=val_col)
                
                txt = str(lbl_cell.value).strip() if lbl_cell.value else ""
                val = clean_value(val_cell.value)
                
                if t_kw.lower() in txt.lower():
                    row[f"{prefix} Total (Sheet)"] = val
                    return

                found = False
                for m_key, m_col in mapping.items():
                    if m_key.lower() in txt.lower():
                        row[m_col] = val
                        found = True
                        break
                
                if not found and "other" in txt.lower():
                    row[f"{prefix} Other Costs"] += val
                    desc = clean_text(txt)
                    if desc: row[f"{prefix} Other Details"].append(desc)
                
                curr += 1

        const_map = {
            "Construction Loan Interest": "Const Loan Interest",
            "Origination Fee": "Const Origination Fee",
            "Credit Enhancement": "Const Credit Enhancement",
            "Bond Premium": "Const Bond Premium",
            "Cost of Issuance": "Const Cost of Issuance",
            "Title & Recording": "Const Title & Recording",
            "Taxes": "Const Taxes", "Insurance": "Const Insurance"
        }
        
        perm_map = {
            "Loan Origination Fee": "Perm Loan Origination Fee",
            "Credit Enhancement": "Perm Credit Enhancement",
            "Title & Recording": "Perm Title & Recording",
            "Taxes": "Perm Taxes", "Insurance": "Perm Insurance"
        }

        get_section_data(["CONSTRUCTION INTEREST & FEES"], "Total Construction Interest & Fees", const_map, "Const")
        get_section_data(["PERMANENT FINANCING"], "Total Permanent Financing Costs", perm_map, "Perm")
        
        row["Const Total (Calculated)"] = sum([row[k] for k in const_map.values()]) + row["Const Other Costs"]
        row["Perm Total (Calculated)"] = sum([row[k] for k in perm_map.values()]) + row["Perm Other Costs"]

        # Hard costs fallback logic
        hc_head = get_header(ws, ["New Construction", "Total New Construction"])
        if hc_head:
             curr = hc_head.row
             for _ in range(20):
                 curr += 1
                 lbl = ws.cell(row=curr, column=hc_head.column).value
                 if lbl and "Total" in str(lbl):
                     val_col = 18 if ws.max_column >= 18 else 3
                     row["Hard Costs"] = clean_value(ws.cell(row=curr, column=val_col).value)
                     break
        
        if row["Hard Costs"] == 0:
            hc_rehab = get_header(ws, ["Rehabilitation", "Total Rehabilitation"])
            if hc_rehab:
                 curr = hc_rehab.row
                 for _ in range(20):
                     curr += 1
                     lbl = ws.cell(row=curr, column=hc_rehab.column).value
                     if lbl and "Total" in str(lbl):
                         val_col = 18 if ws.max_column >= 18 else 3
                         row["Hard Costs"] = clean_value(ws.cell(row=curr, column=val_col).value)
                         row["Flag"].append("Hard Costs Source: Rehab")
                         break
            else:
                row["Flag"].append("Hard Costs Missing")
    else:
        row["Flag"].append("Sources Tab Missing")

    # 2. Units & SF
    app_tab = next((s for s in wb.sheetnames if "Application" in s), None)
    if not app_tab and len(wb.sheetnames) > 0: 
        if "2025" in wb.sheetnames[0]: app_tab = wb.sheetnames[0]

    if app_tab:
        ws_app = wb[app_tab]
        u_kw = [["total", "units"], ["total", "#", "units"], ["unit", "count"], ["total", "residential", "units"]]
        row["Total Units"] = grid_scanner(ws_app, u_kw, 1, 5000)
        
        if 0 < row["Total Units"] < 5:
            row["Flag"].append("Low Unit Count (<5)")
    else:
        row["Flag"].append("App Tab Missing")

    row["Total SF"] = find_sf_in_workbook(wb)
    
    if row["Total SF"] > 0:
        if not app_tab:
            row["Flag"].append("SF Source: Non-App Tab")
    else:
        row["Flag"].append("SF Missing")

    if row["Total Units"] == 0: row["Flag"].append("Units Missing")

    # 3. Validation
    d_const = row["Const Total (Calculated)"] - row["Const Total (Sheet)"]
    if row["Const Total (Sheet)"] == 0 and row["Const Total (Calculated)"] == 0:
        row["Flag"].append("Const Costs Empty")
    elif row["Const Total (Sheet)"] == 0:
        row["Flag"].append("Const Sheet Total Missing")
    elif abs(d_const) > TOLERANCE:
        row["Flag"].append(f"Const Variance (${d_const:+.0f})")

    d_perm = row["Perm Total (Calculated)"] - row["Perm Total (Sheet)"]
    if row["Perm Total (Sheet)"] == 0 and row["Perm Total (Calculated)"] == 0:
        row["Flag"].append("Perm Costs Empty")
    elif row["Perm Total (Sheet)"] == 0:
        row["Flag"].append("Perm Sheet Total Missing")
    elif abs(d_perm) > TOLERANCE:
        row["Flag"].append(f"Perm Variance (${d_perm:+.0f})")

    combined = row["Const Total (Calculated)"] + row["Perm Total (Calculated)"]
    row["Combined Financing Costs"] = combined
    
    if row["Total Units"] > 0:
        row["Financing Cost per Unit"] = combined / row["Total Units"]
        
    if row["Total SF"] > 0:
        row["Financing Cost per SF"] = combined / row["Total SF"]
        
    if row["Hard Costs"] > 0:
        row["% of Hard Costs"] = (combined / row["Hard Costs"]) * 100
    
    return flatten_output(row)

def flatten_output(d):
    d["Const Other Details"] = "; ".join(d["Const Other Details"])
    d["Perm Other Details"] = "; ".join(d["Perm Other Details"])
    d["Flag"] = "; ".join(d["Flag"])
    return d

if __name__ == "__main__":
    if download_files():
        print(f"\nProcessing files in {DATA_DIR}...")
        
        if not os.path.exists(DATA_DIR):
            print("No data directory found.")
        else:
            files = [os.path.join(DATA_DIR, f) for f in os.listdir(DATA_DIR) 
                     if f.lower().endswith(('.xlsx', '.xlsm', '.xls'))]
            
            print(f"Found {len(files)} files.")
            
            records = []
            for i, f in enumerate(files, 1):
                # PROGRESS BAR ADDED HERE
                print(f"[{i}/{len(files)}] Processing {os.path.basename(f)}...")
                records.append(parse_application(f))
            
            df = pd.DataFrame(records)
            
            cols = [
                "File Name", "Flag",
                "Combined Financing Costs", "Financing Cost per Unit", "Financing Cost per SF", "% of Hard Costs",
                "Total Units", "Total SF", "Hard Costs",
                "Const Total (Sheet)", "Const Total (Calculated)", 
                "Const Loan Interest", "Const Origination Fee", "Const Credit Enhancement", "Const Bond Premium", 
                "Const Cost of Issuance", "Const Title & Recording", "Const Taxes", "Const Insurance", 
                "Const Other Costs", "Const Other Details",
                "Perm Total (Sheet)", "Perm Total (Calculated)", 
                "Perm Loan Origination Fee", "Perm Credit Enhancement", "Perm Title & Recording", "Perm Taxes", "Perm Insurance", 
                "Perm Other Costs", "Perm Other Details"
            ]
            
            for c in cols:
                if c not in df.columns: df[c] = 0
                
            df = df[cols]
            df.to_csv(OUTPUT_FILE, index=False)
            print(f"Done. Saved to {OUTPUT_FILE}")